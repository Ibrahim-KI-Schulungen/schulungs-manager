"""
Beleg-Center Modul für Schulungs-Manager
Verwaltet Belege für die Steuererklärung
"""

import streamlit as st
import requests
from datetime import datetime, date, timedelta
from typing import Optional, Dict, List
import io

# Supabase Konfiguration (gleich wie in angebots_pipeline)
def get_supabase_config():
    """Holt Supabase-Konfiguration aus secrets oder env"""
    try:
        url = st.secrets.get("SUPABASE_URL", "")
        key = st.secrets.get("SUPABASE_KEY", "")
    except:
        import os
        url = os.getenv("SUPABASE_URL", "")
        key = os.getenv("SUPABASE_KEY", "")
    return url, key


def supabase_request(method: str, endpoint: str, data: dict = None) -> dict:
    """Führt Supabase REST API Request aus"""
    url, key = get_supabase_config()
    headers = {
        'apikey': key,
        'Authorization': f'Bearer {key}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }

    full_url = f"{url}/rest/v1/{endpoint}"

    try:
        if method == "GET":
            r = requests.get(full_url, headers=headers, timeout=10)
        elif method == "POST":
            r = requests.post(full_url, headers=headers, json=data, timeout=10)
        elif method == "PATCH":
            r = requests.patch(full_url, headers=headers, json=data, timeout=10)
        elif method == "DELETE":
            r = requests.delete(full_url, headers=headers, timeout=10)

        if r.status_code in [200, 201]:
            return {"success": True, "data": r.json() if r.text else []}
        else:
            return {"success": False, "error": f"Status {r.status_code}: {r.text}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_belege(nur_offen: bool = False) -> List[dict]:
    """Holt alle Belege"""
    endpoint = "belege?select=*&order=datum.desc"
    if nur_offen:
        endpoint += "&erledigt=eq.false"

    result = supabase_request("GET", endpoint)
    return result.get("data", []) if result.get("success") else []


def create_beleg(data: dict) -> dict:
    """Erstellt neuen Beleg"""
    return supabase_request("POST", "belege", data)


def update_beleg(beleg_id: str, data: dict) -> dict:
    """Aktualisiert einen Beleg"""
    return supabase_request("PATCH", f"belege?id=eq.{beleg_id}", data)


def delete_beleg(beleg_id: str) -> dict:
    """Löscht einen Beleg"""
    return supabase_request("DELETE", f"belege?id=eq.{beleg_id}")


def get_einstellung(key: str) -> Optional[str]:
    """Holt eine Einstellung"""
    result = supabase_request("GET", f"einstellungen?key=eq.{key}")
    if result.get("success") and result.get("data"):
        return result["data"][0].get("value")
    return None


def set_einstellung(key: str, value: str) -> dict:
    """Setzt eine Einstellung"""
    # Prüfen ob existiert
    existing = get_einstellung(key)
    if existing is not None:
        return supabase_request("PATCH", f"einstellungen?key=eq.{key}", {"value": value})
    else:
        return supabase_request("POST", "einstellungen", {"key": key, "value": value})


def export_belege_excel(belege: List[dict]) -> bytes:
    """Exportiert Belege als Excel-Datei"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Belege"

        # Header
        headers = ["Datum", "Betrag (€)", "Beschreibung", "Kategorie", "Notiz"]
        header_fill = PatternFill(start_color="1f1f33", end_color="1f1f33", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Daten
        for row, beleg in enumerate(belege, 2):
            ws.cell(row=row, column=1, value=beleg.get("datum", ""))
            ws.cell(row=row, column=2, value=float(beleg.get("betrag", 0)))
            ws.cell(row=row, column=3, value=beleg.get("beschreibung", ""))
            ws.cell(row=row, column=4, value=beleg.get("kategorie", ""))
            ws.cell(row=row, column=5, value=beleg.get("notiz", ""))

        # Summenzeile
        sum_row = len(belege) + 2
        ws.cell(row=sum_row, column=1, value="SUMME")
        ws.cell(row=sum_row, column=1).font = Font(bold=True)
        total = sum(float(b.get("betrag", 0)) for b in belege)
        ws.cell(row=sum_row, column=2, value=total)
        ws.cell(row=sum_row, column=2).font = Font(bold=True)

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 30

        # Als Bytes speichern
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except ImportError:
        st.error("openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.")
        return None


KATEGORIEN = ["Software", "Reisekosten", "Büromaterial", "Bewirtung", "Marketing", "Sonstiges"]


def render_beleg_center():
    """Rendert das Beleg-Center"""

    st.markdown("""
    <div class="welcome-header">
        <h2>Beleg-Center</h2>
        <p>Belege für die Steuererklärung verwalten</p>
    </div>
    """, unsafe_allow_html=True)

    # Statistiken
    belege_offen = get_belege(nur_offen=True)
    summe_offen = sum(float(b.get("betrag", 0)) for b in belege_offen)
    letzte_session = get_einstellung("naechste_steuer_session")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Offene Belege", len(belege_offen))
    with col2:
        st.metric("Summe offen", f"{summe_offen:,.2f} €")
    with col3:
        if letzte_session:
            session_date = datetime.strptime(letzte_session, "%Y-%m-%d").date()
            diff = (session_date - date.today()).days
            if diff > 0:
                st.metric("Nächste Session", f"in {diff} Tagen")
            elif diff == 0:
                st.metric("Nächste Session", "Heute!")
            else:
                st.metric("Nächste Session", "Überfällig", delta=f"{abs(diff)} Tage")
        else:
            st.metric("Nächste Session", "Nicht geplant")

    st.markdown("---")

    # Aktionen
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if st.button("➕ Beleg hinzufügen", use_container_width=True, type="primary"):
            st.session_state.show_beleg_modal = True
            st.session_state.edit_beleg = None

    with col2:
        if belege_offen:
            excel_data = export_belege_excel(belege_offen)
            if excel_data:
                st.download_button(
                    "📥 Excel Export",
                    data=excel_data,
                    file_name=f"Belege_Offen_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    with col3:
        if st.button("📅 Session planen", use_container_width=True):
            st.session_state.show_session_modal = True

    with col4:
        if st.session_state.get("selected_belege"):
            if st.button(f"✅ {len(st.session_state.selected_belege)} erledigt", use_container_width=True):
                for beleg_id in st.session_state.selected_belege:
                    update_beleg(beleg_id, {"erledigt": True})
                st.session_state.selected_belege = []
                st.rerun()

    st.markdown("---")

    # Beleg-Liste
    if "selected_belege" not in st.session_state:
        st.session_state.selected_belege = []

    if belege_offen:
        # Header
        col_check, col_datum, col_betrag, col_desc, col_kat, col_actions = st.columns([0.5, 1, 1, 3, 1.5, 1])
        with col_check:
            st.markdown("**☐**")
        with col_datum:
            st.markdown("**Datum**")
        with col_betrag:
            st.markdown("**Betrag**")
        with col_desc:
            st.markdown("**Beschreibung**")
        with col_kat:
            st.markdown("**Kategorie**")
        with col_actions:
            st.markdown("**Aktionen**")

        st.markdown("---")

        # Belege
        for beleg in belege_offen:
            col_check, col_datum, col_betrag, col_desc, col_kat, col_actions = st.columns([0.5, 1, 1, 3, 1.5, 1])

            with col_check:
                is_selected = beleg["id"] in st.session_state.selected_belege
                if st.checkbox("", value=is_selected, key=f"sel_{beleg['id']}", label_visibility="collapsed"):
                    if beleg["id"] not in st.session_state.selected_belege:
                        st.session_state.selected_belege.append(beleg["id"])
                else:
                    if beleg["id"] in st.session_state.selected_belege:
                        st.session_state.selected_belege.remove(beleg["id"])

            with col_datum:
                st.markdown(beleg.get("datum", "-"))

            with col_betrag:
                st.markdown(f"{float(beleg.get('betrag', 0)):,.2f} €")

            with col_desc:
                st.markdown(beleg.get("beschreibung", "-")[:50] + ("..." if len(beleg.get("beschreibung", "")) > 50 else ""))

            with col_kat:
                kat = beleg.get("kategorie", "-")
                kat_colors = {
                    "Software": "#60a5fa",
                    "Reisekosten": "#34d399",
                    "Büromaterial": "#fbbf24",
                    "Bewirtung": "#f472b6",
                    "Marketing": "#a78bfa",
                    "Sonstiges": "#9ca3af"
                }
                color = kat_colors.get(kat, "#666")
                st.markdown(f'<span style="background:{color}; padding:2px 8px; border-radius:4px; font-size:0.8em;">{kat}</span>', unsafe_allow_html=True)

            with col_actions:
                col_a, col_b = st.columns(2)
                with col_a:
                    if st.button("✏️", key=f"edit_b_{beleg['id']}", help="Bearbeiten"):
                        st.session_state.edit_beleg = beleg
                        st.session_state.show_beleg_modal = True
                with col_b:
                    if st.button("🗑️", key=f"del_b_{beleg['id']}", help="Löschen"):
                        delete_beleg(beleg["id"])
                        st.rerun()

        # Summe
        st.markdown("---")
        st.markdown(f"**Summe: {summe_offen:,.2f} €**")

        # Nach Kategorie
        st.markdown("---")
        st.subheader("Nach Kategorie")
        from collections import defaultdict
        by_kat = defaultdict(float)
        for b in belege_offen:
            by_kat[b.get("kategorie", "Sonstiges")] += float(b.get("betrag", 0))

        for kat, summe in sorted(by_kat.items(), key=lambda x: -x[1]):
            st.markdown(f"- **{kat}:** {summe:,.2f} €")

    else:
        st.success("Keine offenen Belege - alles erledigt!")

    # Erledigte Belege (zugeklappt)
    alle_belege = get_belege()
    erledigte = [b for b in alle_belege if b.get("erledigt")]

    if erledigte:
        with st.expander(f"✅ Erledigte Belege ({len(erledigte)})", expanded=False):
            summe_erledigt = sum(float(b.get("betrag", 0)) for b in erledigte)
            st.markdown(f"**Summe:** {summe_erledigt:,.2f} €")
            for b in erledigte[:20]:  # Nur letzte 20
                st.markdown(f"- {b.get('datum')} | {float(b.get('betrag', 0)):,.2f} € | {b.get('beschreibung', '-')[:30]}")

    # Modals
    if st.session_state.get("show_beleg_modal"):
        render_beleg_modal()

    if st.session_state.get("show_session_modal"):
        render_session_modal()


def render_beleg_modal():
    """Modal zum Hinzufügen/Bearbeiten eines Belegs"""

    beleg = st.session_state.get("edit_beleg") or {}
    edit_mode = bool(beleg)

    st.markdown("---")
    st.subheader("✏️ Beleg bearbeiten" if edit_mode else "➕ Neuer Beleg")

    with st.form("beleg_form"):
        col1, col2 = st.columns(2)
        with col1:
            # BUG FIX: Prüfe ob beleg und datum existieren bevor strptime
            default_datum = date.today()
            if beleg and beleg.get("datum"):
                try:
                    default_datum = datetime.strptime(beleg["datum"], "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    pass
            datum = st.date_input("Datum *", value=default_datum)

            # Kategorie mit sicherem Index
            kat_index = 0
            if beleg and beleg.get("kategorie") in KATEGORIEN:
                kat_index = KATEGORIEN.index(beleg["kategorie"])
            kategorie = st.selectbox("Kategorie", options=KATEGORIEN, index=kat_index)
        with col2:
            betrag = st.number_input("Betrag (€) *", value=float(beleg.get("betrag", 0)) if beleg else 0.0, min_value=0.0, step=10.0)

        beschreibung = st.text_input("Beschreibung *", value=beleg.get("beschreibung", "") if beleg else "")
        notiz = st.text_area("Notiz", value=beleg.get("notiz", "") if beleg else "")

        col1, col2, col3 = st.columns(3)
        with col1:
            submitted = st.form_submit_button("💾 Speichern", type="primary", use_container_width=True)
        with col2:
            if edit_mode:
                erledigt = st.form_submit_button("✅ Erledigt", use_container_width=True)
            else:
                erledigt = False
        with col3:
            cancel = st.form_submit_button("❌ Abbrechen", use_container_width=True)

        if submitted:
            if not beschreibung or not betrag:
                st.error("Bitte alle Pflichtfelder ausfüllen")
            else:
                data = {
                    "datum": datum.isoformat(),
                    "betrag": betrag,
                    "beschreibung": beschreibung,
                    "kategorie": kategorie,
                    "notiz": notiz or None
                }

                if edit_mode:
                    result = update_beleg(beleg["id"], data)
                else:
                    result = create_beleg(data)

                if result.get("success"):
                    st.success("Gespeichert!")
                    st.session_state.show_beleg_modal = False
                    st.session_state.edit_beleg = None
                    st.rerun()
                else:
                    st.error(f"Fehler: {result.get('error')}")

        if erledigt:
            update_beleg(beleg["id"], {"erledigt": True})
            st.session_state.show_beleg_modal = False
            st.session_state.edit_beleg = None
            st.rerun()

        if cancel:
            st.session_state.show_beleg_modal = False
            st.session_state.edit_beleg = None
            st.rerun()


def render_session_modal():
    """Modal zum Planen der Steuer-Session"""

    st.markdown("---")
    st.subheader("📅 Steuer-Session planen")

    aktuelle = get_einstellung("naechste_steuer_session")

    with st.form("session_form"):
        default_date = date.today() + timedelta(days=14)
        if aktuelle:
            try:
                default_date = datetime.strptime(aktuelle, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass

        session_datum = st.date_input("Datum der nächsten Session", value=default_date)

        # Email-Erinnerung (Coming Soon)
        st.checkbox("📧 Erinnerung 1 Tag vorher per Email", value=False, disabled=True, help="Coming Soon")

        col1, col2 = st.columns(2)
        with col1:
            save_btn = st.form_submit_button("💾 Speichern", type="primary", use_container_width=True)
        with col2:
            cancel_btn = st.form_submit_button("❌ Abbrechen", use_container_width=True)

    # Logik NACH dem Form-Block (wichtig für Streamlit!)
    if save_btn:
        result = set_einstellung("naechste_steuer_session", session_datum.isoformat())
        if result.get("success"):
            st.success(f"✅ Session für {session_datum.strftime('%d.%m.%Y')} geplant!")
            st.session_state.show_session_modal = False
            st.rerun()
        else:
            st.error(f"Fehler: {result.get('error', 'Unbekannt')}")

    if cancel_btn:
        st.session_state.show_session_modal = False
        st.rerun()


def get_beleg_notifications() -> List[dict]:
    """Gibt Benachrichtigungen für Dashboard zurück"""
    notifications = []

    # Steuer-Session bald?
    session_str = get_einstellung("naechste_steuer_session")
    if session_str:
        session_date = datetime.strptime(session_str, "%Y-%m-%d").date()
        diff = (session_date - date.today()).days
        if diff <= 7 and diff >= 0:
            belege_offen = len(get_belege(nur_offen=True))
            notifications.append({
                "type": "warning",
                "icon": "🟡",
                "text": f"Steuer-Session in {diff} Tagen — {belege_offen} offene Belege",
                "link": "beleg_center"
            })
        elif diff < 0:
            notifications.append({
                "type": "error",
                "icon": "🔴",
                "text": f"Steuer-Session überfällig!",
                "link": "beleg_center"
            })

    return notifications
